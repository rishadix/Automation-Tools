#!/usr/bin/env python
# coding: utf-8

# Tool Script for IDQ Kit is below:

# In[1]:


## Updated Script without fs-automation drive access:

import pandas as pd
import os
import numpy as np
import tkinter as tk
from tkinter import filedialog,messagebox,simpledialog
import datetime
from datetime import datetime, date
import sys
import openpyxl
from openpyxl import load_workbook
import time

name=os.getlogin()
messagebox.showinfo(message='Choose Input File')
path = filedialog.askopenfilename()
asas = pd.read_excel(path, sheet_name=None)
aas = asas['Input']

## 3P Flat File:
aas_3p = aas[['asin','ptd','attribute','value','unit']][aas['deprecated_hack_merchant_id.value'].isnull()==True].reset_index(drop=True)
aas_3p.rename(columns={'value':'attribute_value'}, inplace=True)
main_input = aas_3p
# flat_3P()
master_file = pd.read_excel(r"C:\Users\rishadix\Desktop\Rishabh\3P_Raw_Flat_file\Master_3P_Flat_Tool.xlsx",sheet_name=None)
master_file['Master'].rename(columns = {'unit':'attribute_unit'},inplace=True)
master = master_file['Master']
try:    
    start =datetime.now()
#         ab=pd.read_excel(path,sheet_name=None)
#         main_input=ab['Sheet1']
#         main_input
    aa=pd.merge(main_input,master[['ptd','attribute','attribute_unit']].drop_duplicates(),on=['ptd','attribute'],how='left')
    aa
    aa.rename(columns={'attribute_value':'value'},inplace=True)
    wq1=aa[['asin','attribute_unit','unit']][aa['attribute_unit'].isnull()!=True]
    wq1.rename(columns={'attribute_unit':'attribute','unit':'value'},inplace=True)
    wq1
    wq=pd.concat([aa[['asin','attribute','value',]],wq1])
    wq.sort_values(by=['asin','attribute'])
    wq.reset_index(drop=True,inplace=True)
    wq.fillna("",inplace=True)

    header = master_file['Raw_flat'].iloc[1]                               # To Optimize the Code
    raw = master_file['Raw_flat'].rename(columns=header)                   # To Optimize the Code

    raw = raw.iloc[2:]
    raw = raw.reset_index(drop=True)
    raw['item_sku'] = wq['asin']
    raw.fillna('',axis=1, inplace=True)
#-------------------------------------    
#     mapping_dict = dict(zip(wq['attribute'], wq['value']))
# #     raw.update(pd.DataFrame(mapping_dict, index=wq.index))
#     for col in raw.columns:
#         if col in mapping_dict:
#             raw[col] = mapping_dict[col]
##---------------------------------------------------------------------------------------------------------------------------
    for i in range(len(wq)):
        for col in raw.columns:
            if col == wq['attribute'][i]:
                raw.at[i, col] = wq['value'][i]
##---------------------------------------------------------------------------------------------------------------------------
    raw.drop(columns='feed_product_type', inplace=True)
    raw = raw.pivot_table(index='item_sku',aggfunc=lambda x: ''.join(x.astype('str')), sort=False).reset_index()
    
    ## IDQ Tool Kit Indexing Issue
#     raw.drop(columns='feed_product_type', inplace=True)                        
#     raw = raw.pivot_table(index='item_sku', aggfunc=lambda x: ''.join(x.astype('str')), sort=False, dropna=False).reset_index()
#     raw.reset_index(inplace=True)    

    raw['external_product_id'] = raw['item_sku']
    raw[['external_product_id_type', 'update_delete']] = ['ASIN', 'PartialUpdate']
    new_data = ''
    raw.insert(0, 'feed_product_type', new_data)
    raw = raw.transpose().reset_index() 
    raw
##---------------------------------------------------------------------------------------------------------------------------

    Raw_file = pd.read_excel(r"C:\Users\rishadix\Desktop\Rishabh\3P_Raw_Flat_file\Master_3P_Flat_Tool.xlsx",sheet_name='Raw_flat', nrows=2, header=None, na_values=['']).transpose()         # To concatinate Raw Flat_file
    Auto_Flat_File = pd.concat([Raw_file, raw], axis=1, ignore_index=True).transpose()
    Auto_Flat_File.fillna('',axis=1, inplace=True)
except KeyError as e:
##     pass
    print(f"New Attribute found: <<{e}>> \nPlease make sure that name of the Attribute is Correct. \nIf it's a New Attribute, Kindly raise a SIM to get this added in Sample Flat File list\n")

dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":","-")
end = datetime.now()                                                   # To check Time Duration
Time_Taken = end-start                                                          # To check Time Duration
print(f'3P Task Completed !! Time_Taken: {(Time_Taken.seconds)} seconds')
# Auto_Flat_File.to_csv('C:/Users/rishadix/Downloads/3P_Flat_File_'+dt+'.csv', header=0, index=False)
# To update the Task Details:
# wb = openpyxl.load_workbook(r"//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Biswa/Log_file.xlsx")
# ws = wb['RAC_Flat_File_Tool']

# ab = pd.DataFrame({'Alias': 'rishadix', 'Count_Of_Records': aa.count(), 'Time_taken (in Seconds)': Time_Taken.seconds, 'Date': [date.today()]})
# # print(ab)
# for row in dataframe_to_rows(ab, index=False, header=False):
#     ws.append(row)
# wb.save(r"//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Biswa/Log_file.xlsx")
# Auto_Flat_File.to_csv(path[:path.rfind('/')]+'3P_Flat_File_'+dt+'.csv', header=0, index=False)

##---------------------------------------------------------------------------------------------------------------------------------

## RA Flat File:
aas_ra = aas[['asin','ptd','attribute','value','tag','unit']][aas['deprecated_hack_merchant_id.value'].isnull()!=True].reset_index(drop=True)
aas_ra.rename(columns={'tag':'language_tag'}, inplace=True)
main_input = aas_ra
# flat_RA()
master_file=pd.read_excel(r"C:\Users\rishadix\Desktop\Rishabh\Retail_Flat_File\Master_RA_Flat_Tool.xlsx",sheet_name=None)
master_file['Master'].rename(columns={'value':'attribute_value','unit':'attribute_unit'},inplace=True)
master=master_file['Master']
start_time =datetime.now()
try:
#         ab=pd.read_excel(path,sheet_name=None)
#         main_input=ab['Input']
#         main_input
    wq=pd.merge(main_input,master[['ptd','attribute','attribute_value','attribute_unit']].drop_duplicates(),on=['ptd','attribute'],how='left')

    # Changes made due to Blank Attribute Error:
    wq.drop(columns='attribute',inplace=True)
    wq.rename(columns={'attribute_value':'attribute'},inplace=True)
    wq[wq['attribute']=='fit_type#1.value']            # To cross check
    wq2=wq[['asin','attribute_unit','unit']][wq['attribute_unit'].isnull()!=True]
    wq2.rename(columns={'attribute_unit':'attribute','unit':'value'},inplace=True)
    wq2
    wq1=pd.concat([wq[['asin','attribute','value',]],wq2])
    wq1.sort_values(by=['asin','attribute'])
    wq1.reset_index(drop=True,inplace=True)
    wq1.fillna("",inplace=True)

##    bqe=asas['BQE'].drop_duplicates()                                               ## For Separate RA Input File
    bqe = aas[['asin','deprecated_hack_merchant_id.value']].drop_duplicates()         ## For Combined RA & 3P Input File
    bqe = bqe.rename(columns={'deprecated_hack_merchant_id.value':'contributor_id', 'asin':'ASIN'})
    bqe=pd.merge(bqe,master_file['Mapping'],on='contributor_id')
    bqe
    new_piv=wq1.pivot_table(index='asin',columns='attribute',values='value',aggfunc=lambda x: ','.join(x.astype('str')))
    new_piv.insert(loc=0,column='ASIN',value=new_piv.index)
    new_piv

    # Attribute Error Message:
    if (new_piv.columns == '').any():                                    
        error = new_piv[['ASIN','']][new_piv[''].isnull()!=True]
        print(f"New Attribute found for these ASINs/values:\n{error}\nPlease make sure that name of the Attribute is Correct. \nIf it's a New Attribute, Kindly raise a SIM to get this added in Sample Flat File list\n")
##         return            # It works on RA Tool Script
#             sys.exit()          # Only for Jupyter

    bqe=pd.merge(bqe,new_piv,on='ASIN')
    bqe.loc[-1]=bqe.columns
    bqe.index = bqe.index + 1 
    bqe= bqe.sort_index()
    bqe.columns=master_file['Raw_flat'].columns[0:len(bqe.columns)]
#     bqe= bqe.reset_index().drop(columns='index')
    bqe.loc[0,'version=1.0.0']='sku'            # On Neha's Request
    bqe
    bqe.columns=[""  if i.startswith('Unnamed') else i for i in bqe.columns]
except KeyError as e:
##     pass
    print(e)

end_time =datetime.now()
Time_Taken = end_time-start_time
print(f"RA Task Completed !! Time_Taken: {(end_time-start_time).seconds} seconds")

## To update the Task Log Details:
# ab = pd.DataFrame({'Alias': x, 'Count_Of_Records': main_input['asin'].count(), 'Time_taken (in Seconds)': Time_Taken, 'Date': [date.today()]})
# book = load_workbook(r'//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Rishabh/Retail_Flat_File/Tool_Log_File.xlsx')
# if 'RA_Flat_Test' in book.sheetnames:
#     sheet1 = book['RA_Flat_Test']
# else:
#     sheet1 = book.create_sheet('RA_Flat_Test')
# last_row = sheet1.max_row
# for index, row in ab.iterrows():
#     for i, value in enumerate(row):
#         sheet1.cell(row=last_row + index + 1, column=i + 1, value=value)
# book.save('//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Rishabh/Retail_Flat_File/Tool_Log_File.xlsx')

dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":","-")
# bqe.to_csv(path[:path.rfind('/')]+'RA_Flat_File_'+dt+'.csv', header=0, index=False)
# bqe.to_csv('C:/Users/'+name+'/Downloads/'+'RA_Flat_File_'+dt+'.csv', header=0, index=False)
# Auto_Flat_File.to_csv('C:/Users/'+name+'/Downloads/'+'3P_Flat_File_'+dt+'.csv', header=0, index=False)

bqe
Auto_Flat_File


# In[6]:


import pandas as pd
import os
import numpy as np
import tkinter as tk
from tkinter import filedialog,messagebox,simpledialog
import datetime
from datetime import datetime, date
import sys
import openpyxl
from openpyxl import load_workbook
import time

name=os.getlogin()
# messagebox.showinfo(message='Choose Input File')
# path = filedialog.askopenfilename()
asas = pd.read_excel(path, sheet_name=None)
aas = asas['Input']

## 3P Flat File:
aas_3p = aas[['asin','ptd','attribute','value','unit']][aas['deprecated_hack_merchant_id.value'].isnull()==True].reset_index(drop=True)
aas_3p.rename(columns={'value':'attribute_value'}, inplace=True)
main_input = aas_3p
# flat_3P()
master_file = pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\3P_Raw_Flat_file\Master_3P_Flat_Tool.xlsx",sheet_name=None)
master_file['Master'].rename(columns = {'unit':'attribute_unit'},inplace=True)
master = master_file['Master']
try:    
    start =datetime.now()
#         ab=pd.read_excel(path,sheet_name=None)
#         main_input=ab['Sheet1']
#         main_input
    aa=pd.merge(main_input,master[['ptd','attribute','attribute_unit']].drop_duplicates(),on=['ptd','attribute'],how='left')
    aa
    aa.rename(columns={'attribute_value':'value'},inplace=True)
    wq1=aa[['asin','attribute_unit','unit']][aa['attribute_unit'].isnull()!=True]
    wq1.rename(columns={'attribute_unit':'attribute','unit':'value'},inplace=True)
    wq1
    wq=pd.concat([aa[['asin','attribute','value',]],wq1])
    wq.sort_values(by=['asin','attribute'])
    wq.reset_index(drop=True,inplace=True)
    wq.fillna("",inplace=True)

    header = master_file['Raw_flat'].iloc[1]                               # To Optimize the Code
    raw = master_file['Raw_flat'].rename(columns=header)                   # To Optimize the Code

    raw = raw.iloc[2:]
    raw = raw.reset_index(drop=True)
    raw['item_sku'] = wq['asin']
    raw.fillna('',axis=1, inplace=True)
#-------------------------------------    
#     mapping_dict = dict(zip(wq['attribute'], wq['value']))
# #     raw.update(pd.DataFrame(mapping_dict, index=wq.index))
#     for col in raw.columns:
#         if col in mapping_dict:
#             raw[col] = mapping_dict[col]
##---------------------------------------------------------------------------------------------------------------------------
    for i in range(len(wq)):
        for col in raw.columns:
            if col == wq['attribute'][i]:
                raw.at[i, col] = wq['value'][i]
##---------------------------------------------------------------------------------------------------------------------------
    raw.drop(columns='feed_product_type', inplace=True)
    raw = raw.pivot_table(index='item_sku',aggfunc=lambda x: ''.join(x.astype('str')), sort=False).reset_index()
    
    ## IDQ Tool Kit Indexing Issue
#     raw.drop(columns='feed_product_type', inplace=True)                        
#     raw = raw.pivot_table(index='item_sku', aggfunc=lambda x: ''.join(x.astype('str')), sort=False, dropna=False).reset_index()
#     raw.reset_index(inplace=True)    

    raw['external_product_id'] = raw['item_sku']
    raw[['external_product_id_type', 'update_delete']] = ['ASIN', 'PartialUpdate']
    new_data = ''
    raw.insert(0, 'feed_product_type', new_data)
    raw = raw.transpose().reset_index() 
    raw
##---------------------------------------------------------------------------------------------------------------------------

    Raw_file = pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\3P_Raw_Flat_file\Master_3P_Flat_Tool.xlsx",sheet_name='Raw_flat', nrows=2, header=None, na_values=['']).transpose()         # To concatinate Raw Flat_file
    Auto_Flat_File = pd.concat([Raw_file, raw], axis=1, ignore_index=True).transpose()
    Auto_Flat_File.fillna('',axis=1, inplace=True)
except KeyError as e:
##     pass
    print(f"New Attribute found: <<{e}>> \nPlease make sure that name of the Attribute is Correct. \nIf it's a New Attribute, Kindly raise a SIM to get this added in Sample Flat File list\n")

dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":","-")
end = datetime.now()                                                   # To check Time Duration
Time_Taken = end-start                                                          # To check Time Duration
print(f'3P Task Completed !! Time_Taken: {(Time_Taken.seconds)} seconds')
# Auto_Flat_File.to_csv('C:/Users/rishadix/Downloads/3P_Flat_File_'+dt+'.csv', header=0, index=False)
# To update the Task Details:
# wb = openpyxl.load_workbook(r"//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Biswa/Log_file.xlsx")
# ws = wb['RAC_Flat_File_Tool']

# ab = pd.DataFrame({'Alias': 'rishadix', 'Count_Of_Records': aa.count(), 'Time_taken (in Seconds)': Time_Taken.seconds, 'Date': [date.today()]})
# # print(ab)
# for row in dataframe_to_rows(ab, index=False, header=False):
#     ws.append(row)
# wb.save(r"//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Biswa/Log_file.xlsx")
# Auto_Flat_File.to_csv(path[:path.rfind('/')]+'3P_Flat_File_'+dt+'.csv', header=0, index=False)

##---------------------------------------------------------------------------------------------------------------------------------

## RA Flat File:
aas_ra = aas[['asin','ptd','attribute','value','tag','unit']][aas['deprecated_hack_merchant_id.value'].isnull()!=True].reset_index(drop=True)
aas_ra.rename(columns={'tag':'language_tag'}, inplace=True)
main_input = aas_ra
# flat_RA()
master_file=pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\Retail_Flat_File\Master_RA_Flat_Tool.xlsx",sheet_name=None)
master_file['Master'].rename(columns={'value':'attribute_value','unit':'attribute_unit'},inplace=True)
master=master_file['Master']
start_time =datetime.now()
try:
#         ab=pd.read_excel(path,sheet_name=None)
#         main_input=ab['Input']
#         main_input
    wq=pd.merge(main_input,master[['ptd','attribute','attribute_value','attribute_unit']].drop_duplicates(),on=['ptd','attribute'],how='left')

    # Changes made due to Blank Attribute Error:
    wq.drop(columns='attribute',inplace=True)
    wq.rename(columns={'attribute_value':'attribute'},inplace=True)
    wq[wq['attribute']=='fit_type#1.value']            # To cross check
    wq2=wq[['asin','attribute_unit','unit']][wq['attribute_unit'].isnull()!=True]
    wq2.rename(columns={'attribute_unit':'attribute','unit':'value'},inplace=True)
    wq2
    wq1=pd.concat([wq[['asin','attribute','value',]],wq2])
    wq1.sort_values(by=['asin','attribute'])
    wq1.reset_index(drop=True,inplace=True)
    wq1.fillna("",inplace=True)

##    bqe=asas['BQE'].drop_duplicates()                                               ## For Separate RA Input File
    bqe = aas[['asin','deprecated_hack_merchant_id.value']].drop_duplicates()         ## For Combined RA & 3P Input File
    bqe = bqe.rename(columns={'deprecated_hack_merchant_id.value':'contributor_id', 'asin':'ASIN'})
    bqe=pd.merge(bqe,master_file['Mapping'],on='contributor_id')
    bqe
    new_piv=wq1.pivot_table(index='asin',columns='attribute',values='value',aggfunc=lambda x: ','.join(x.astype('str')))
    new_piv.insert(loc=0,column='ASIN',value=new_piv.index)
    new_piv

    # Attribute Error Message:
    if (new_piv.columns == '').any():                                    
        error = new_piv[['ASIN','']][new_piv[''].isnull()!=True]
        print(f"New Attribute found for these ASINs/values:\n{error}\nPlease make sure that name of the Attribute is Correct. \nIf it's a New Attribute, Kindly raise a SIM to get this added in Sample Flat File list\n")
##         return            # It works on RA Tool Script
#             sys.exit()          # Only for Jupyter

    bqe=pd.merge(bqe,new_piv,on='ASIN')
    bqe.loc[-1]=bqe.columns
    bqe.index = bqe.index + 1 
    bqe= bqe.sort_index()
    bqe.columns=master_file['Raw_flat'].columns[0:len(bqe.columns)]
#     bqe= bqe.reset_index().drop(columns='index')
    bqe.loc[0,'version=1.0.0']='sku'            # On Neha's Request
    bqe
    bqe.columns=[""  if i.startswith('Unnamed') else i for i in bqe.columns]
except KeyError as e:
##     pass
    print(e)

end_time =datetime.now()
Time_Taken = end_time-start_time
print(f"RA Task Completed !! Time_Taken: {(end_time-start_time).seconds} seconds")

## To update the Task Log Details:
# ab = pd.DataFrame({'Alias': x, 'Count_Of_Records': main_input['asin'].count(), 'Time_taken (in Seconds)': Time_Taken, 'Date': [date.today()]})
# book = load_workbook(r'//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Rishabh/Retail_Flat_File/Tool_Log_File.xlsx')
# if 'RA_Flat_Test' in book.sheetnames:
#     sheet1 = book['RA_Flat_Test']
# else:
#     sheet1 = book.create_sheet('RA_Flat_Test')
# last_row = sheet1.max_row
# for index, row in ab.iterrows():
#     for i, value in enumerate(row):
#         sheet1.cell(row=last_row + index + 1, column=i + 1, value=value)
# book.save('//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Rishabh/Retail_Flat_File/Tool_Log_File.xlsx')

dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":","-")
# bqe.to_csv(path[:path.rfind('/')]+'RA_Flat_File_'+dt+'.csv', header=0, index=False)
# bqe.to_csv('C:/Users/'+name+'/Downloads/'+'RA_Flat_File_'+dt+'.csv', header=0, index=False)
# Auto_Flat_File.to_csv('C:/Users/'+name+'/Downloads/'+'3P_Flat_File_'+dt+'.csv', header=0, index=False)

bqe
Auto_Flat_File


# In[5]:


# Auto_Flat_File.to_csv('C:/Users/'+name+'/Downloads/'+'3P_Flat_Test_'+dt+'.csv', header=0, index=False)
raw.update()


# ### Tool Kit Script of 'Combined RA & 3P Flat File Tool':

# In[172]:


import pandas as pd
import os
import numpy as np
import tkinter as tk
from tkinter import filedialog,messagebox,simpledialog
import datetime
from datetime import datetime, date
import sys
import openpyxl
from openpyxl import load_workbook
import time

def RA_3P_flat(x):
    print(f"Hey {x}!! Work in Progress...")
    name=os.getlogin()
    messagebox.showinfo(message='Choose Input File')
    path = filedialog.askopenfilename()
    asas = pd.read_excel(path, sheet_name=None)
    aas = asas['Input']

    ## 3P Flat File:
    aas_3p = aas[['asin','ptd','attribute','value','unit']][aas['deprecated_hack_merchant_id.value'].isnull()==True].reset_index(drop=True)
    aas_3p.rename(columns={'value':'attribute_value'}, inplace=True)
    main_input = aas_3p

    master_file = pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\3P_Raw_Flat_file\Master_3P_Flat_Tool.xlsx",sheet_name=None)
    master_file['Master'].rename(columns = {'unit':'attribute_unit'},inplace=True)
    master = master_file['Master']
    try:    
        start =datetime.now()
        aa=pd.merge(main_input,master[['ptd','attribute','attribute_unit']].drop_duplicates(),on=['ptd','attribute'],how='left')
        aa
        aa.rename(columns={'attribute_value':'value'},inplace=True)
        wq1=aa[['asin','attribute_unit','unit']][aa['attribute_unit'].isnull()!=True]
        wq1.rename(columns={'attribute_unit':'attribute','unit':'value'},inplace=True)
        wq1
        wq=pd.concat([aa[['asin','attribute','value',]],wq1])
        wq.sort_values(by=['asin','attribute'])
        wq.reset_index(drop=True,inplace=True)
        wq.fillna("",inplace=True)

        header = master_file['Raw_flat'].iloc[1]                               # To Optimize the Code
        raw = master_file['Raw_flat'].rename(columns=header)                   # To Optimize the Code

        raw = raw.iloc[2:]
        raw = raw.reset_index(drop=True)
        raw['item_sku'] = wq['asin']
        raw.fillna('',axis=1, inplace=True)

        for i in range(len(wq)):
            for col in raw.columns:
                if col == wq['attribute'][i]:
                    raw.at[i, col] = wq['value'][i]
    ##---------------------------------------------------------------------------------------------------------------------------
        raw.drop(columns='feed_product_type', inplace=True)
        raw = raw.pivot_table(index='item_sku',aggfunc=lambda x: ''.join(x.astype('str')), sort=False).reset_index()
        raw['external_product_id'] = raw['item_sku']
        raw[['external_product_id_type', 'update_delete']] = ['ASIN', 'PartialUpdate']
        new_data = ''
        raw.insert(0, 'feed_product_type', new_data)
        raw = raw.transpose().reset_index() 
        raw
    ##---------------------------------------------------------------------------------------------------------------------------

        Raw_file = pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\3P_Raw_Flat_file\Master_3P_Flat_Tool.xlsx",sheet_name='Raw_flat', nrows=2, header=None, na_values=['']).transpose()         # To concatinate Raw Flat_file
        Auto_Flat_File = pd.concat([Raw_file, raw], axis=1, ignore_index=True).transpose()
        Auto_Flat_File.fillna('',axis=1, inplace=True)
    except KeyError as e:
    ##     pass
        print(f"New Attribute found: <<{e}>> \nPlease make sure that name of the Attribute is Correct. \nIf it's a New Attribute, Kindly raise a SIM to get this added in Sample Flat File list\n")

    dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":","-")
    end = datetime.now()                                                   # To check Time Duration
    Time_Taken = end-start                                                          # To check Time Duration
    print(f'3P Task Completed !! Time_Taken: {(Time_Taken.seconds)} seconds')
    # messagebox.showinfo(message= 'Done')

    ## To update the Task Details:
    book = r"//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Biswa/Log_file.xlsx"
    ax = pd.DataFrame({'Alias':x, 'Count_Of_Records': aa.count(), 'Time_taken (in Seconds)': Time_Taken.seconds, 'Date': dt},index=[0])
    #ax=pd.DataFrame(ab,index=[0])
    fla_file=pd.read_excel(book,sheet_name='RAC_Flat_File_Tool')
    fla_file=pd.concat([fla_file,ax],ignore_index=True)
    with pd.ExcelWriter(book,engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        fla_file.to_excel(writer,sheet_name='RAC_Flat_File_Tool',index=False)

    ##---------------------------------------------------------------------------------------------------------------------------------

    ## RA Flat File:
    aas_ra = aas[['asin','ptd','attribute','value','tag','unit']][aas['deprecated_hack_merchant_id.value'].isnull()!=True].reset_index(drop=True)
    aas_ra.rename(columns={'tag':'language_tag'}, inplace=True)
    main_input = aas_ra

    master_file=pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\Retail_Flat_File\Master_RA_Flat_Tool.xlsx",sheet_name=None)
    master_file['Master'].rename(columns={'value':'attribute_value','unit':'attribute_unit'},inplace=True)
    master=master_file['Master']
    start_time =datetime.now()
    try:
        wq=pd.merge(main_input,master[['ptd','attribute','attribute_value','attribute_unit']].drop_duplicates(),on=['ptd','attribute'],how='left')

        # Changes made due to Blank Attribute Error:
        wq.drop(columns='attribute',inplace=True)
        wq.rename(columns={'attribute_value':'attribute'},inplace=True)
        wq[wq['attribute']=='fit_type#1.value']            # To cross check
        wq2=wq[['asin','attribute_unit','unit']][wq['attribute_unit'].isnull()!=True]
        wq2.rename(columns={'attribute_unit':'attribute','unit':'value'},inplace=True)
        wq2
        wq1=pd.concat([wq[['asin','attribute','value',]],wq2])
        wq1.sort_values(by=['asin','attribute'])
        wq1.reset_index(drop=True,inplace=True)
        wq1.fillna("",inplace=True)

    ##    bqe=asas['BQE'].drop_duplicates()                                   ## For Separate RA Input File
        bqe = aas[['asin','deprecated_hack_merchant_id.value']]             ## For Combined RA & 3P Input File
        bqe=bqe.rename(columns={'deprecated_hack_merchant_id.value':'contributor_id', 'asin':'ASIN'})
        bqe=pd.merge(bqe,master_file['Mapping'],on='contributor_id')
        bqe
        new_piv=wq1.pivot_table(index='asin',columns='attribute',values='value',aggfunc=lambda x: ','.join(x.astype('str')))
        new_piv.insert(loc=0,column='ASIN',value=new_piv.index)
        new_piv

        # Attribute Error Message:
        if (new_piv.columns == '').any():                                    
            error = new_piv[['ASIN','']][new_piv[''].isnull()!=True]
            print(f"New Attribute found for these ASINs/values:\n{error} \nPlease make sure that name of the Attribute is Correct. \nIf it's a New Attribute, Kindly raise a SIM to get this added in Sample Flat File list\n")
            return            # It works only for IDQ_Tool Script

        bqe=pd.merge(bqe,new_piv,on='ASIN')
        bqe.loc[-1]=bqe.columns
        bqe.index = bqe.index + 1 
        bqe= bqe.sort_index()
        bqe.columns=master_file['Raw_flat'].columns[0:len(bqe.columns)]
        bqe.loc[0,'version=1.0.0']='sku'            # On Neha's Request
        bqe
        bqe.columns=[""  if i.startswith('Unnamed') else i for i in bqe.columns]
    except KeyError as e:
    ##     pass
        print(e)

    end_time =datetime.now()
    Time_Taken = end_time-start_time
    print(f"RA Task Completed !! Time_Taken: {(end_time-start_time).seconds} seconds")

        # To update the Task Log Details:
    ab = pd.DataFrame({'Alias': x, 'Count_Of_Records': main_input['asin'].count(), 'Time_taken (in Seconds)': Time_Taken, 'Date': [date.today()]})
    book = load_workbook(r'//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Rishabh/Retail_Flat_File/Tool_Log_File.xlsx')
    if 'RA_Flat_Test' in book.sheetnames:
        sheet1 = book['RA_Flat_Test']
    else:
        sheet1 = book.create_sheet('RA_Flat_Test')

    last_row = sheet1.max_row

    for index, row in ab.iterrows():
        for i, value in enumerate(row):
            sheet1.cell(row=last_row + index + 1, column=i + 1, value=value)

    book.save('//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Rishabh/Retail_Flat_File/Tool_Log_File.xlsx')

    dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":","-")
    # bqe.to_csv(path[:path.rfind('/')]+'RA_Flat_File_'+dt+'.csv', header=0, index=False)
    bqe.to_csv('C:/Users/'+name+'/Downloads/'+'RA_Flat_File_'+dt+'.csv', header=0, index=False)
    Auto_Flat_File.to_csv('C:/Users/'+name+'/Downloads/'+'3P_Flat_File_'+dt+'.csv', header=0, index=False)
    messagebox.showinfo(message= 'Done')
    print("Task Completed")
    # bqe


# In[170]:


name=os.getlogin()
name

# aa[aas['deprecated_hack_merchant_id.value']!='']


# In[97]:


## 3P Script:
import pandas as pd
import os
import numpy as np
import tkinter as tk
from tkinter import filedialog,messagebox,simpledialog
import datetime
from datetime import datetime

# messagebox.showinfo(message='Choose File')
# path = filedialog.askopenfilename()

def flat_3P():
#     messagebox.showinfo(message='Choose Input File')
#     path = filedialog.askopenfilename()
    master_file = pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\3P_Raw_Flat_file\Master_3P_Flat_Tool.xlsx",sheet_name=None)
    master_file['Master'].rename(columns = {'unit':'attribute_unit'},inplace=True)
    master = master_file['Master']
        ## master[master['attribute_value']=='fit_type#1.value']                                # To cross check
    try:    
        start =datetime.now()
#         ab=pd.read_excel(path,sheet_name=None)
#         main_input=ab['Sheet1']
#         main_input
        aa=pd.merge(main_input,master[['ptd','attribute','attribute_unit']].drop_duplicates(),on=['ptd','attribute'],how='left')
        aa
        aa.rename(columns={'attribute_value':'value'},inplace=True)
        wq1=aa[['asin','attribute_unit','unit']][aa['attribute_unit'].isnull()!=True]
        wq1.rename(columns={'attribute_unit':'attribute','unit':'value'},inplace=True)
        wq1
        wq=pd.concat([aa[['asin','attribute','value',]],wq1])
        wq.sort_values(by=['asin','attribute'])
        wq.reset_index(drop=True,inplace=True)
        wq.fillna("",inplace=True)

        header = master_file['Raw_flat'].iloc[1]                               # To Optimize the Code
        raw = master_file['Raw_flat'].rename(columns=header)                   # To Optimize the Code

        raw = raw.iloc[2:]
        raw = raw.reset_index(drop=True)
        raw['item_sku'] = wq['asin']
        raw.fillna('',axis=1, inplace=True)

        for i in range(len(wq)):
            for col in raw.columns:
                if col == wq['attribute'][i]:
                    raw.at[i, col] = wq['value'][i]
    ##---------------------------------------------------------------------------------------------------------------------------
        raw.drop(columns='feed_product_type', inplace=True)
        raw = raw.pivot_table(index='item_sku',aggfunc=lambda x: ''.join(x.astype('str')), sort=False).reset_index()
        raw['external_product_id'] = raw['item_sku']
        raw[['external_product_id_type', 'update_delete']] = ['ASIN', 'PartialUpdate']
        new_data = ''
        raw.insert(0, 'feed_product_type', new_data)
        raw = raw.transpose().reset_index() 
        raw
    ##---------------------------------------------------------------------------------------------------------------------------

        Raw_file = pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\3P_Raw_Flat_file\Master_3P_Flat_Tool.xlsx",sheet_name='Raw_flat', nrows=2, header=None, na_values=['']).transpose()         # To concatinate Raw Flat_file
        Auto_Flat_File = pd.concat([Raw_file, raw], axis=1, ignore_index=True).transpose()
        Auto_Flat_File.fillna('',axis=1, inplace=True)
    except KeyError as e:
    ##     pass
        print(f"New Attribute found: <<{e}>> \nPlease make sure that name of the Attribute is Correct. \nIf it's a New Attribute, Kindly raise a SIM to get this added in Sample Flat File list\n")

    dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":","-")
    end = datetime.now()                                                   # To check Time Duration
    Time_Taken = end-start                                                          # To check Time Duration
    print(f'Task Completed !! Time_Taken: {(Time_Taken.seconds)} seconds')
    Auto_Flat_File.to_csv('C:/Users/rishadix/Downloads/3P_Flat_File_'+dt+'.csv', header=0, index=False)
    # # To update the Task Details:
    # wb = openpyxl.load_workbook(r"//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Biswa/Log_file.xlsx")
    # ws = wb['RAC_Flat_File_Tool']

    # ab = pd.DataFrame({'Alias': 'rishadix', 'Count_Of_Records': aa.count(), 'Time_taken (in Seconds)': Time_Taken.seconds, 'Date': [date.today()]})
    # # print(ab)
    # for row in dataframe_to_rows(ab, index=False, header=False):
    #     ws.append(row)
    # wb.save(r"//ant/dept-as/blr2-Groupdata1/FS-AutomationTechnologies/Biswa/Log_file.xlsx")
#     Auto_Flat_File.to_csv(path[:path.rfind('/')]+'3P_Flat_File_'+dt+'.csv', header=0, index=False)

# Auto_Flat_File


# In[98]:


## Retail Optimized Script:
import pandas as pd
import os
import numpy as np
import tkinter as tk
from tkinter import filedialog,messagebox,simpledialog
from datetime import datetime
import sys
# messagebox.showinfo(message='Choose File')
# path = filedialog.askopenfilename()

def flat_RA():
#     messagebox.showinfo(message='Choose File')
#     path = filedialog.askopenfilename()
    master_file=pd.read_excel(r"\\ant\dept-as\blr2-Groupdata1\FS-AutomationTechnologies\Rishabh\Retail_Flat_File\Master_RA_Flat_Tool.xlsx",sheet_name=None)
    master_file['Master'].rename(columns={'value':'attribute_value','unit':'attribute_unit'},inplace=True)
    master=master_file['Master']
    ## master[master['attribute_value']=='fit_type#1.value']                                # To cross check
    start_time =datetime.now()
    try:
#         ab=pd.read_excel(path,sheet_name=None)
#         main_input=ab['Input']
#         main_input
        wq=pd.merge(main_input,master[['ptd','attribute','attribute_value','attribute_unit']].drop_duplicates(),on=['ptd','attribute'],how='left')

        # ---------------------------------------------------------------------------------------------
        # Changes made due to Blank Attribute Error:
        wq.drop(columns='attribute',inplace=True)
        wq.rename(columns={'attribute_value':'attribute'},inplace=True)
        wq[wq['attribute']=='fit_type#1.value']            # To cross check
        # ---------------------------------------------------------------------------------------------
        # wq.replace(list(wq['attribute']),list(wq['attribute_value']),inplace=True)         # removed post error
        ## wq[wq['attribute']=='fit_type#1.value']
        # wq.drop(columns='attribute_value',inplace=True)                                    # removed post error
        # ---------------------------------------------------------------------------------------------
        wq2=wq[['asin','attribute_unit','unit']][wq['attribute_unit'].isnull()!=True]
        wq2.rename(columns={'attribute_unit':'attribute','unit':'value'},inplace=True)
        wq2
        wq1=pd.concat([wq[['asin','attribute','value',]],wq2])
        wq1.sort_values(by=['asin','attribute'])
        wq1.reset_index(drop=True,inplace=True)
        wq1.fillna("",inplace=True)
        ## wq1[wq1['attribute']=='fit_type#1.value']                                    # To cross check
        bqe=ab['BQE'].drop_duplicates()
        bqe.rename(columns={'deprecated_hack_merchant_id.value':'contributor_id'},inplace=True)
        bqe=pd.merge(bqe,master_file['Mapping'],on='contributor_id')
        bqe
        new_piv=wq1.pivot_table(index='asin',columns='attribute',values='value',aggfunc=lambda x: ','.join(x.astype('str')))
        new_piv.insert(loc=0,column='ASIN',value=new_piv.index)
        new_piv
    ##----------------------------------------------------------------------------------------------    
        # Attribute Error Message:
        if (new_piv.columns == '').any():                                    
            error = new_piv[['ASIN','']][new_piv[''].isnull()!=True]
            print(error)
    ##         return            # It works on RA Tool Script
#             sys.exit()          # Only for Jupyter
    ##----------------------------------------------------------------------------------------------    

        bqe=pd.merge(bqe,new_piv,on='ASIN')
        bqe.loc[-1]=bqe.columns
        bqe.index = bqe.index + 1 
        bqe= bqe.sort_index()
        bqe.columns=master_file['Raw_flat'].columns[0:len(bqe.columns)]
        bqe.loc[0,'version=1.0.0']='sku'            # On Neha's Request
        bqe
        bqe.columns=[""  if i.startswith('Unnamed') else i for i in bqe.columns]
    except KeyError as e:
    ##     pass
        print(f"New Attribute found: <<{e}>> \nPlease make sure that name of the Attribute is Correct. \nIf it's a New Attribute, Kindly raise a SIM to get this added in Sample Flat File list\n")

    end_time =datetime.now()
    Time_Taken = end_time-start_time
    print(f"Task Completed !! Time_Taken: {(end_time-start_time).seconds} seconds")
    dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":","-")
    bqe.to_csv(path[:path.rfind('/')]+'RA_Flat_File_'+dt+'.csv', header=0, index=False)
    # bqe
# new_piv


# In[163]:


# new_piv.loc[1]=
# b=aas[['asin','contributor_id']]
aas=aas[['asin','tag','ptd','attribute','value','unit','deprecated_hack_merchant_id.value']]
aas['value'].str.upper()
# aas.rename(columns={'deprecated_hack_merchant_id.value':'contributor_id'},inplace=True)
# bqe=pd.merge(aas[['asin','deprecated_hack_merchant_id.value']],master_file['Mapping'],on='contributor_id')
# bqe
# new_piv = pd.merge(new_piv, aas[['asin','deprecated_hack_merchant_id.value']], on='asin', how='left')
# new_piv['deprecated_hack_merchant_id.value'].loc[2]
# new_piv.drop['']
# new_piv
# bqe


# In[21]:


len(raw)


# In[ ]:




